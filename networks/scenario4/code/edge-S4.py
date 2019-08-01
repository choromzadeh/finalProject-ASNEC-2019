from openpyxl import Workbook, load_workbook
import pandas

workBookDataSet = load_workbook("Airports_reduction.xlsx")

sheetDataSet = workBookDataSet['Airports_reduction']

edgesWorkBook = Workbook()
edgesSheet = edgesWorkBook.active

SourceEdges = []
TargetEdges = []
WeightEdges = [1000]

endNodes=[]
print(WeightEdges[10])

# for x in range(1,8000):
#     for y in range(x,8000):
#         if str(sheetDataSet['A' + str(x)].value) + str(sheetDataSet['B' + str(x)].value) == str(sheetDataSet['A' + str(y)].value) + str(sheetDataSet['B' + str(y)].value):
#
#
# for x in range(len(SourceEdges)):
#     edgesSheet['A' + str(x+1)] = SourceEdges[x]
#     edgesSheet['B' + str(x+1)] = TargetEdges[x]
#     edgesSheet['C' + str(x+1)] = WeightEdges[x]



# print(edgesSheet['A1'].value + edgesSheet['B1'].value)
# print(SourceEdges[0] + TargetEdges[0] + str(WeightEdges[0]))

edgesWorkBook.save('S1-edges.xlsx')
# SourceEdges = []
# TargetEdges = []

# for x in sourceEdges:
#     for y in nodes:
#         if x.value == y.value :
#            newSourceEdges.append(nodes.index(y))
#
# for x in TargetEdges:
#     for y in nodes:
#         if x.value == y.value :
#            newTargetEdges.append(nodes.index(y))
#
#
# for x in range(2,56):
#     sheetEdges.cell(row=x,column=4,value=newSourceEdges[x-2])
#     sheetEdges.cell(row=x, column=5, value=newTargetEdges[x - 2])




# workBookEdges.save('newEdge2.xlsx')

# print(sheetEdges['D1'].value)
# print(sheetEdges['D2'].value)
# print(sheetEdges['D3'].value)
# print(sheetEdges['D4'].value)
# print(sheetEdges['D5'].value)
# print(sheetEdges['D6'].value)
# print(sheetEdges['D55'].value)
# print('..................')
#
# print(sheetEdges['E1'].value)
# print(sheetEdges['E2'].value)
# print(sheetEdges['E3'].value)






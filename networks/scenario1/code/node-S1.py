from openpyxl import Workbook, load_workbook


workBookDataSet = load_workbook("flight-reduction(1000).xlsx")

sheetDataSet = workBookDataSet['dataSets2']

nodesWorkBook = Workbook()
nodesSheet = nodesWorkBook.active

primaryNodes = sheetDataSet['E']

endNodes=[]

for x in range(1,12000):
    nodesSheet['A' + str(x)] = sheetDataSet['E' + str(x)].value



nodesWorkBook.save('S1-nodes.xlsx')
SourceEdges = []
TargetEdges = []

# for x in sourceEdges:
#     for y in nodes:
#         if x.value == y.value :
#            newSourceEdges.append(nodes.index(y))
#
# for x in TargetEdges:
#     for y in nodes:
#         if x.value == y.value :
#            newTargetEdges.append(nodes.index(y))
#
#
# for x in range(2,56):
#     sheetEdges.cell(row=x,column=4,value=newSourceEdges[x-2])
#     sheetEdges.cell(row=x, column=5, value=newTargetEdges[x - 2])




# workBookEdges.save('newEdge2.xlsx')

# print(sheetEdges['D1'].value)
# print(sheetEdges['D2'].value)
# print(sheetEdges['D3'].value)
# print(sheetEdges['D4'].value)
# print(sheetEdges['D5'].value)
# print(sheetEdges['D6'].value)
# print(sheetEdges['D55'].value)
# print('..................')
#
# print(sheetEdges['E1'].value)
# print(sheetEdges['E2'].value)
# print(sheetEdges['E3'].value)






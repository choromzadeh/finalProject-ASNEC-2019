from openpyxl import Workbook, load_workbook
import pandas

workBookDataSet = load_workbook("flight-reduction(1000).xlsx")

sheetDataSet = workBookDataSet['dataSets2']

edgesWorkBook = Workbook()
edgesSheet = edgesWorkBook.active

SourceEdges = []
TargetEdges = []
WeightEdges = []

endNodes=[]


for x in range(1,8000):
    for y in range(x,8000):
        if str(sheetDataSet['H' + str(x)].value) + str(sheetDataSet['I' + str(x)].value) == str(sheetDataSet['H' + str(y)].value) + str(sheetDataSet['I' + str(y)].value):
                if str(sheetDataSet['E' + str(x)].value) != str(sheetDataSet['E' + str(y)].value):
                        # print(str(sheetDataSet['E' + str(x)].value + sheetDataSet['E' + str(y)].value ))
                        SourceEdges.append(sheetDataSet['E' + str(x)].value)
                        TargetEdges.append(sheetDataSet['E' + str(y)].value)
                        if type(sheetDataSet['L' + str(x)].value) == int and type(sheetDataSet['L' + str(y)].value) == int:
                            # print(abs(abs(sheetDataSet['L' + str(x)].value) - abs(sheetDataSet['L' + str(y)].value)))
                            WeightEdges.append(abs(abs(sheetDataSet['L' + str(x)].value) - abs(sheetDataSet['L' + str(y)].value)))
                        else:
                            WeightEdges.append(0)

for x in range(len(SourceEdges)):
    edgesSheet['A' + str(x+1)] = SourceEdges[x]
    edgesSheet['B' + str(x+1)] = TargetEdges[x]
    edgesSheet['C' + str(x+1)] = WeightEdges[x]



# print(edgesSheet['A1'].value + edgesSheet['B1'].value)
# print(SourceEdges[0] + TargetEdges[0] + str(WeightEdges[0]))

edgesWorkBook.save('S1-edges.xlsx')
# SourceEdges = []
# TargetEdges = []

# for x in sourceEdges:
#     for y in nodes:
#         if x.value == y.value :
#            newSourceEdges.append(nodes.index(y))
#
# for x in TargetEdges:
#     for y in nodes:
#         if x.value == y.value :
#            newTargetEdges.append(nodes.index(y))
#
#
# for x in range(2,56):
#     sheetEdges.cell(row=x,column=4,value=newSourceEdges[x-2])
#     sheetEdges.cell(row=x, column=5, value=newTargetEdges[x - 2])




# workBookEdges.save('newEdge2.xlsx')

# print(sheetEdges['D1'].value)
# print(sheetEdges['D2'].value)
# print(sheetEdges['D3'].value)
# print(sheetEdges['D4'].value)
# print(sheetEdges['D5'].value)
# print(sheetEdges['D6'].value)
# print(sheetEdges['D55'].value)
# print('..................')
#
# print(sheetEdges['E1'].value)
# print(sheetEdges['E2'].value)
# print(sheetEdges['E3'].value)





